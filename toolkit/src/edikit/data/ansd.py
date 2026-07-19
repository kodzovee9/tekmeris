"""Reader for ANSD (Senegal) supply-use tables — the Tableau des Ressources
et des Emplois (TRE).

Targets the workbook ``TRE_en_valeur_2014-2023.xlsx``: one sheet per year,
each a full TRE in a fixed French layout (products A00..ZB0 as rows,
branches as columns). Blocks are located by their labels, not by fixed row
numbers, because the layout shifts by a row or two across years — the same
"verify by content, never by position" discipline the rest of the toolkit
follows.

Sheet layout (one year):
  - a supply block ("Ressources en produits"): for each product, the make
    matrix across branches, the basic->purchaser valuation bridge (trade
    and transport margins, non-deductible VAT, product/export/import taxes,
    subsidies), domestic output, a cif/fob adjustment, and imports;
  - a use block ("Emploi des produits"): for each product, intermediate
    use across branches and final demand (exports, household consumption
    split own/marketed, government and NPISH consumption, gross fixed
    capital formation, inventories, valuables);
  - a value-added block by branch: gross value added, compensation of
    employees (gross wages plus actual and imputed social contributions),
    taxes and subsidies on production, and gross operating surplus / mixed
    income.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl

# product / branch codes: A00..Z00 plus ZA0, ZB0 (28 in total)
_CODE = re.compile(r"^(?:[A-Z]00|Z[AB]0)$")


def _norm(s) -> str:
    """Accent- and case-insensitive form for matching French labels that
    vary in accent encoding across ANSD vintages."""
    return "".join(c for c in unicodedata.normalize("NFKD", str(s))
                   if not unicodedata.combining(c)).lower()


@dataclass(frozen=True)
class Finding:
    check: str
    subject: str
    detail: str


@dataclass
class SenegalTRE:
    """One year's TRE: make and use matrices plus the valuation, final-demand
    and value-added vectors. All values in the workbook's units (CFA millions,
    current prices)."""

    year: int
    products: list[str]                     # product codes in row order
    branches: list[str]                     # branch codes in column order
    labels: dict[str, str] = field(default_factory=dict)  # code -> description
    make: dict[tuple[str, str], float] = field(default_factory=dict)  # (product, branch)
    use: dict[tuple[str, str], float] = field(default_factory=dict)   # (product, branch)
    # supply-side vectors, by product
    supply_purchaser: dict[str, float] = field(default_factory=dict)
    trade_margin: dict[str, float] = field(default_factory=dict)
    transport_margin: dict[str, float] = field(default_factory=dict)
    vat: dict[str, float] = field(default_factory=dict)
    subsidies_on_products: dict[str, float] = field(default_factory=dict)
    other_product_taxes: dict[str, float] = field(default_factory=dict)
    export_taxes: dict[str, float] = field(default_factory=dict)
    import_taxes: dict[str, float] = field(default_factory=dict)
    supply_basic: dict[str, float] = field(default_factory=dict)
    output: dict[str, float] = field(default_factory=dict)
    cif_fob: dict[str, float] = field(default_factory=dict)
    imports: dict[str, float] = field(default_factory=dict)
    # use-side final demand, by product
    use_purchaser: dict[str, float] = field(default_factory=dict)
    exports: dict[str, float] = field(default_factory=dict)
    cons_households: dict[str, float] = field(default_factory=dict)
    cons_autoconsumption: dict[str, float] = field(default_factory=dict)
    cons_marketed: dict[str, float] = field(default_factory=dict)
    cons_government: dict[str, float] = field(default_factory=dict)
    cons_npish: dict[str, float] = field(default_factory=dict)
    gfcf: dict[str, float] = field(default_factory=dict)
    inventories: dict[str, float] = field(default_factory=dict)
    valuables: dict[str, float] = field(default_factory=dict)
    # value-added block, by branch
    value_added: dict[str, float] = field(default_factory=dict)
    compensation: dict[str, float] = field(default_factory=dict)
    wages: dict[str, float] = field(default_factory=dict)
    social_contributions: dict[str, float] = field(default_factory=dict)
    production_taxes: dict[str, float] = field(default_factory=dict)
    production_subsidies: dict[str, float] = field(default_factory=dict)
    operating_surplus: dict[str, float] = field(default_factory=dict)

    def output_by_branch(self) -> dict[str, float]:
        out = {b: 0.0 for b in self.branches}
        for (_, b), v in self.make.items():
            out[b] += v
        return out

    def intermediate_by_branch(self) -> dict[str, float]:
        out = {b: 0.0 for b in self.branches}
        for (_, b), v in self.use.items():
            out[b] += v
        return out


def _num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def read_tre(path: str, year: int) -> SenegalTRE:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[str(year)]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    def cell(r0: int, c0: int):  # 0-indexed
        return grid[r0][c0] if 0 <= r0 < len(grid) and c0 < len(grid[r0]) else None

    # --- locate the branch-code header row: the row carrying the most codes
    # across its columns (product rows carry a code in column 1 only) ---
    branch_row, branch_cols = None, []
    for i, row in enumerate(grid[:14]):
        codes = [j for j, v in enumerate(row)
                 if v is not None and _CODE.match(str(v).strip())]
        if len(codes) > len(branch_cols):
            branch_row, branch_cols = i, codes
    if branch_row is None or len(branch_cols) < 2:
        raise ValueError(f"{path} [{year}]: branch header row not found")
    branches = [str(cell(branch_row, c)).strip() for c in branch_cols]
    # branch descriptions live a couple of rows above, same columns
    desc_row = next((i for i in range(branch_row) if
                     isinstance(cell(i, branch_cols[0]), str)
                     and len(str(cell(i, branch_cols[0])).strip()) > 4), branch_row)

    # --- the "Emploi des produits" divider splits supply (above) from use ---
    divider = next((i for i, row in enumerate(grid)
                    if any(isinstance(v, str) and "emploi des produits" in _norm(v)
                           for v in row)), None)
    if divider is None:
        raise ValueError(f"{path} [{year}]: use block ('Emploi') not found")

    # --- column map: within a sheet the supply valuation bridge is a fixed
    # block (columns 3..11), then the branch run, then output/cif-fob/imports ---
    last_branch = branch_cols[-1]
    C = {
        "purchaser": 2, "trade": 3, "transport": 4, "vat": 5, "subv": 6,
        "otax": 7, "xtax": 8, "mtax": 9, "basic": 10,
        "output": last_branch + 1, "cif": last_branch + 2, "imports": last_branch + 3,
    }

    tre = SenegalTRE(year=year, products=[], branches=branches)
    for c, b in zip(branch_cols, branches):
        tre.labels[b] = str(cell(desc_row, c) or "").strip()

    # --- SUPPLY block: product rows between the header and the divider ---
    for i in range(branch_row + 1, divider):
        code = str(cell(i, 0) or "").strip()
        if not _CODE.match(code):
            continue
        tre.products.append(code)
        tre.labels.setdefault(code, str(cell(i, 1) or "").strip())
        for c, b in zip(branch_cols, branches):
            v = _num(cell(i, c))
            if v:
                tre.make[(code, b)] = v
        tre.supply_purchaser[code] = _num(cell(i, C["purchaser"]))
        tre.trade_margin[code] = _num(cell(i, C["trade"]))
        tre.transport_margin[code] = _num(cell(i, C["transport"]))
        tre.vat[code] = _num(cell(i, C["vat"]))
        tre.subsidies_on_products[code] = _num(cell(i, C["subv"]))
        tre.other_product_taxes[code] = _num(cell(i, C["otax"]))
        tre.export_taxes[code] = _num(cell(i, C["xtax"]))
        tre.import_taxes[code] = _num(cell(i, C["mtax"]))
        tre.supply_basic[code] = _num(cell(i, C["basic"]))
        tre.output[code] = _num(cell(i, C["output"]))
        tre.cif_fob[code] = _num(cell(i, C["cif"]))
        tre.imports[code] = _num(cell(i, C["imports"]))

    # --- USE block: final-demand columns are labelled after the branch run ---
    def fd_col(*subs):
        for hr in (divider, divider + 1, divider + 2):
            for j in range(last_branch + 1, ws.max_column):
                v = cell(hr, j)
                if isinstance(v, str) and any(_norm(s) in _norm(v) for s in subs):
                    return j
        return None

    c_exports = fd_col("Expor")
    c_cons = fd_col("Consommation finale", "Cons. Fin")
    c_hh = fd_col("Sous-total", "Ménages")
    c_auto = fd_col("Autocons")
    c_mkt = fd_col("Commerc")
    c_gov = fd_col("Adminis", "trations")
    c_npish = fd_col("ISBL")
    c_gfcf = fd_col("capital fixe")
    c_stock = fd_col("Variations des stocks")
    c_val = fd_col("objets de valeur")

    for i in range(divider + 1, len(grid)):
        code = str(cell(i, 0) or "").strip()
        if not _CODE.match(code):
            continue
        for c, b in zip(branch_cols, branches):
            v = _num(cell(i, c))
            if v:
                tre.use[(code, b)] = v
        tre.use_purchaser[code] = _num(cell(i, C["purchaser"]))
        tre.exports[code] = _num(cell(i, c_exports)) if c_exports else 0.0
        tre.cons_households[code] = _num(cell(i, c_hh)) if c_hh else 0.0
        tre.cons_autoconsumption[code] = _num(cell(i, c_auto)) if c_auto else 0.0
        tre.cons_marketed[code] = _num(cell(i, c_mkt)) if c_mkt else 0.0
        tre.cons_government[code] = _num(cell(i, c_gov)) if c_gov else 0.0
        tre.cons_npish[code] = _num(cell(i, c_npish)) if c_npish else 0.0
        tre.gfcf[code] = _num(cell(i, c_gfcf)) if c_gfcf else 0.0
        tre.inventories[code] = _num(cell(i, c_stock)) if c_stock else 0.0
        tre.valuables[code] = _num(cell(i, c_val)) if c_val else 0.0

    # --- VALUE-ADDED block: rows located by label, values across branches ---
    def va_row(*subs):
        for i in range(divider, len(grid)):
            lab = " ".join(str(cell(i, k) or "") for k in (0, 1))
            if any(_norm(s) in _norm(lab) for s in subs):
                return i
        return None

    def read_va(row_i, target):
        if row_i is None:
            return
        for c, b in zip(branch_cols, branches):
            target[b] = _num(cell(row_i, c))

    read_va(va_row("Valeur ajoutée brute", "PIB"), tre.value_added)
    read_va(va_row("Rémunération des salariés"), tre.compensation)
    read_va(va_row("Salaires bruts"), tre.wages)
    # actual + imputed social contributions summed
    soc = {}
    for sub in ("Contributions sociales effectives", "Contributions sociales imputées"):
        r = va_row(sub)
        if r is not None:
            for c, b in zip(branch_cols, branches):
                soc[b] = soc.get(b, 0.0) + _num(cell(r, c))
    tre.social_contributions = soc
    read_va(va_row("Impôts sur la production"), tre.production_taxes)
    read_va(va_row("Subventions sur la production"), tre.production_subsidies)
    read_va(va_row("Excédent brut d'exploitation"), tre.operating_surplus)
    return tre


def check_tre(tre: SenegalTRE, tol: float = 1.0) -> list[Finding]:
    """Enforce the TRE's accounting identities; return the failures."""
    findings: list[Finding] = []

    # 1. supply at basic prices = domestic output + imports + cif/fob adjustment
    for p in tre.products:
        lhs = tre.supply_basic[p]
        rhs = tre.output[p] + tre.imports[p] + tre.cif_fob[p]
        if abs(lhs - rhs) > tol:
            findings.append(Finding("supply-origin", p,
                f"basic {lhs:,.0f} vs output+imports+cif {rhs:,.0f}"))

    # 2. total supply at purchaser prices = total use at purchaser prices
    for p in tre.products:
        s, u = tre.supply_purchaser[p], tre.use_purchaser[p]
        if abs(s - u) > tol:
            findings.append(Finding("supply-use-balance", p,
                f"supply {s:,.0f} vs use {u:,.0f}"))

    # 3. branch output = intermediate consumption + value added
    out_b = tre.output_by_branch()
    int_b = tre.intermediate_by_branch()
    for b in tre.branches:
        lhs = out_b[b]
        rhs = int_b[b] + tre.value_added.get(b, 0.0)
        if abs(lhs - rhs) > max(tol, 1e-4 * abs(lhs)):
            findings.append(Finding("output-identity", b,
                f"output {lhs:,.0f} vs intermediate+VA {rhs:,.0f}"))

    # 4. value added = compensation + production taxes - subsidies + operating surplus
    for b in tre.branches:
        va = tre.value_added.get(b, 0.0)
        rhs = (tre.compensation.get(b, 0.0) + tre.production_taxes.get(b, 0.0)
               + tre.production_subsidies.get(b, 0.0)
               + tre.operating_surplus.get(b, 0.0))
        if abs(va - rhs) > max(tol, 1e-4 * abs(va)):
            findings.append(Finding("va-decomposition", b,
                f"VA {va:,.0f} vs comp+taxes+GOS {rhs:,.0f}"))

    return findings


# --- institutional-sector accounts (TCEI) -------------------------------------

# sector sheet -> canonical institution code (matching the ANSD MCS accounts)
_TCEI_SECTORS = {"0S1001": "SNF", "0S1002": "SF", "0S1003": "APU",
                 "0S1004": "Men", "0S1005": "ISBL", "0S2": "Rdm"}
# "Compte de ..." block -> short account name (order: check 'ajuste' before plain).
# The last two entries are the rest-of-the-world sheet's (0S2) external accounts,
# which fold the primary-allocation and secondary-distribution transactions into
# one "external current" block rather than splitting them the way the domestic
# sector sheets do.
_TCEI_ACCOUNTS = [
    ("production", "production"),
    ("exploitation", "generation"),
    ("affectation des revenus primaires", "primary"),
    ("distribution secondaire", "secondary"),
    ("revenu en nature", "inkind"),
    ("revenu disponible ajuste", "adjusted"),
    ("utilisation du revenu disponible", "useincome"),
    ("capital", "capital"),
    ("exterieur des biens et services", "extgoods"),
    ("exterieur des revenus primaires", "extcurrent"),
]
_TX = re.compile(r"^[A-Z]\.\d")           # SNA codes: D.1, P.11, B.2/B.3, K.2, ...


@dataclass
class TCEI:
    """One year's institutional-sector accounts, keyed by
    (sector, account, direction, transaction-code)."""

    year: int
    flows: dict[tuple[str, str, str, str], float] = field(default_factory=dict)

    def get(self, sector: str, account: str, direction: str, code: str,
            default: float = 0.0) -> float:
        return self.flows.get((sector, account, direction, code), default)


def read_tcei(path: str, year: int) -> TCEI:
    """Read the ANSD institutional-sector accounts for one year. Each sector
    sheet is the SNA sequence of accounts; the reader tracks the account
    block and the resources/emplois side so a transaction that recurs across
    accounts (e.g. D.1 paid in generation, received in primary allocation)
    stays distinct."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = TCEI(year=year)
    for sheet, sector in _TCEI_SECTORS.items():
        if sheet not in wb.sheetnames:
            continue
        grid = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        # year column: the header cell equal to the requested year
        ycol = None
        for r in grid[:8]:
            for j, v in enumerate(r):
                if isinstance(v, (int, float)) and int(v) == year:
                    ycol = j
            if ycol is not None:
                break
        if ycol is None:
            continue
        # Transaction codes are in column B (index 1) and the Ressources/Emplois
        # markers in column C (index 2). Account-block headers ("Compte de ...")
        # sit in column C on the domestic and rest-of-the-world sheets but in
        # column B on the NPISH (0S1005) sheet, so look for them in either.
        account, direction = None, None
        for r in grid:
            code = str(r[1]).strip() if len(r) > 1 and r[1] else ""
            colb = _norm(str(r[1]).strip()) if len(r) > 1 and r[1] else ""
            ctx = _norm(str(r[2]).strip()) if len(r) > 2 and r[2] else ""
            htext = ctx if "compte" in ctx else (colb if "compte" in colb else "")
            hdr = next((short for sub, short in _TCEI_ACCOUNTS
                        if htext and _norm(sub) in htext), None)
            if hdr:
                account, direction = hdr, None
            elif ctx.startswith("ressources"):
                direction = "resources"
            elif ctx.startswith("emplois"):
                direction = "emplois"
            elif _TX.match(code) and account and direction:
                v = r[ycol] if ycol < len(r) else None
                if isinstance(v, (int, float)):
                    out.flows[(sector, account, direction, code)] = float(v)
    return out


# --- the official ANSD SAM (MCS) as an audit target --------------------------

@dataclass
class MCS:
    """One year of the official ANSD social accounting matrix."""

    year: int
    accounts: list[str]                     # 65 account codes in order
    labels: dict[str, str] = field(default_factory=dict)
    cells: dict[tuple[str, str], float] = field(default_factory=dict)  # (row, col)


def read_mcs(path: str, year: int) -> MCS:
    """Read a year sheet of the official MCS. Row account codes are in column B;
    the square matrix follows in the same account order (columns start one to
    the right of the labels), with a trailing TOTAL row and column."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    grid = [list(r) for r in wb[str(year)].iter_rows(values_only=True)]
    # row index of each account code (column B = index 1), skipping TOTAL
    rows: list[tuple[str, int]] = []
    for i, r in enumerate(grid):
        code = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        if code and code != "TOTAL":
            rows.append((code, i))
    accounts = [c for c, _ in rows]
    # row labels sit in column B (index 1); the square matrix begins in the
    # next column and runs in the same account order, so column of the k-th
    # account is (label column + 1) + k
    label_col = 1
    data_c0 = label_col + 1
    mcs = MCS(year=year, accounts=accounts)
    col_of = {code: data_c0 + k for k, code in enumerate(accounts)}
    for code, ri in rows:
        for ccode, cj in col_of.items():
            v = grid[ri][cj] if cj < len(grid[ri]) else None
            if isinstance(v, (int, float)) and v:
                mcs.cells[(code, ccode)] = float(v)
    return mcs
