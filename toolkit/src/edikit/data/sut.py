"""Reader for Statistics South Africa supply and use table workbooks.

Targets the Report 04-04-03 layout (e.g. "Supply and use tables 2018 - 2019.xlsx"):
sheets "Industry List", "Product List", "Supply Table <year>", "Use Table <year>".

Supply table layout: row 1 holds column headers; product rows start at row 4 and
carry (code, SIC, description) in columns 1-3; adjustment columns 4-7 hold total
supply at purchasers' prices, taxes less subsidies, trade/transport margins, and
total supply at basic prices; industry columns are headed I1..In. A totals row
follows the product block (identified by an empty description cell).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

_PRODUCT_CODE = re.compile(r"^P\d+$")
_INDUSTRY_CODE = re.compile(r"^I\d+$")


@dataclass(frozen=True)
class Industry:
    code: str
    description: str
    sic_detail: str = ""
    sic_published: str = ""


@dataclass(frozen=True)
class Product:
    code: str
    sic: str
    description: str


@dataclass
class SupplyTable:
    """One year's supply table: products x industries plus adjustment vectors."""

    year: int
    products: list[Product]
    industries: list[str]  # industry codes in column order
    # matrix[(product_code, industry_code)] -> value (R'million); zeros omitted
    matrix: dict[tuple[str, str], float] = field(default_factory=dict)
    total_supply_purchasers: dict[str, float] = field(default_factory=dict)
    taxes_less_subsidies: dict[str, float] = field(default_factory=dict)
    margins: dict[str, float] = field(default_factory=dict)
    total_supply_basic: dict[str, float] = field(default_factory=dict)
    # column totals as embedded in the sheet's totals row, by industry code
    embedded_output_totals: dict[str, float] = field(default_factory=dict)

    def output_by_industry(self) -> dict[str, float]:
        """Recompute industry output (column sums) from the matrix."""
        out: dict[str, float] = {code: 0.0 for code in self.industries}
        for (_, icode), v in self.matrix.items():
            out[icode] += v
        return out


@dataclass(frozen=True)
class Finding:
    """A failed or noteworthy check, with the numbers that triggered it."""

    check: str
    subject: str
    detail: str


def read_industry_list(path: str) -> list[Industry]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = list(wb["Industry List"].iter_rows(values_only=True))
    industries = []
    for r in rows[1:]:
        code = str(r[0] or "").strip()
        if _INDUSTRY_CODE.match(code):
            industries.append(
                Industry(
                    code=code,
                    description=str(r[1] or "").strip(),
                    sic_detail=str(r[2] or "").strip(),
                    sic_published=str(r[3] or "").strip(),
                )
            )
    return industries


def read_supply_table(path: str, year: int) -> SupplyTable:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = wb[f"Supply Table {year}"]
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]

    industry_cols: dict[int, str] = {}
    for j, h in enumerate(header):
        code = str(h or "").strip()
        if _INDUSTRY_CODE.match(code):
            industry_cols[j] = code

    products: list[Product] = []
    table = SupplyTable(
        year=year,
        products=products,
        industries=list(industry_cols.values()),
    )

    for r in rows[3:]:
        code = str(r[0] or "").strip()
        desc = str(r[2] or "").strip()
        if not _PRODUCT_CODE.match(code):
            continue
        if not desc:
            # totals row reuses a product-like code but has no description
            for j, icode in industry_cols.items():
                if isinstance(r[j], (int, float)):
                    table.embedded_output_totals[icode] = float(r[j])
            continue
        products.append(Product(code=code, sic=str(r[1] or "").strip(), description=desc))
        if isinstance(r[3], (int, float)):
            table.total_supply_purchasers[code] = float(r[3])
        if isinstance(r[4], (int, float)):
            table.taxes_less_subsidies[code] = float(r[4])
        if isinstance(r[5], (int, float)):
            table.margins[code] = float(r[5])
        if isinstance(r[6], (int, float)):
            table.total_supply_basic[code] = float(r[6])
        for j, icode in industry_cols.items():
            v = r[j]
            if isinstance(v, (int, float)) and v != 0:
                table.matrix[(code, icode)] = float(v)

    return table


def check_supply_table(table: SupplyTable, tolerance: float = 0.5) -> list[Finding]:
    """Internal consistency checks. `tolerance` is in R'million (absolute).

    Note: total supply at basic prices also includes imports and c.i.f./f.o.b.
    adjustments, which are not part of the domestic output matrix, so the row
    identity checked here is purchasers' = basic + taxes + margins only.
    """
    findings: list[Finding] = []
    for p in table.products:
        purch = table.total_supply_purchasers.get(p.code)
        basic = table.total_supply_basic.get(p.code)
        taxes = table.taxes_less_subsidies.get(p.code, 0.0)
        marg = table.margins.get(p.code, 0.0)
        if purch is None or basic is None:
            findings.append(Finding("row-identity", p.code, "missing totals cell"))
            continue
        gap = purch - (basic + taxes + marg)
        if abs(gap) > tolerance:
            findings.append(
                Finding(
                    "row-identity",
                    p.code,
                    f"purchasers' {purch:,.1f} != basic {basic:,.1f} "
                    f"+ taxes {taxes:,.1f} + margins {marg:,.1f} (gap {gap:,.1f})",
                )
            )
    recomputed = table.output_by_industry()
    for icode, embedded in table.embedded_output_totals.items():
        gap = recomputed.get(icode, 0.0) - embedded
        if abs(gap) > tolerance:
            findings.append(
                Finding(
                    "column-total",
                    icode,
                    f"recomputed {recomputed.get(icode, 0.0):,.1f} != embedded {embedded:,.1f}",
                )
            )
    return findings


@dataclass
class UseTable:
    """One year's use table: intermediate consumption, final demand, value added.

    Verified identities for this layout (encoded in check_use_table):
    per product, total supply at purchasers' = intermediate + final demand
    (including the Residual column); per industry, output = intermediate + B1,
    and B1 = D1 + D2/3 + B2/3.
    """

    year: int
    products: list[Product]
    industries: list[str]
    fd_categories: list[str]
    va_components: dict[str, str] = field(default_factory=dict)  # code -> label
    # intermediate[(product_code, industry_code)] -> value; zeros omitted
    intermediate: dict[tuple[str, str], float] = field(default_factory=dict)
    final_demand: dict[tuple[str, str], float] = field(default_factory=dict)
    value_added: dict[tuple[str, str], float] = field(default_factory=dict)  # (va_code, industry)
    total_supply_purchasers: dict[str, float] = field(default_factory=dict)
    embedded_intermediate_by_product: dict[str, float] = field(default_factory=dict)
    embedded_intermediate_by_industry: dict[str, float] = field(default_factory=dict)
    embedded_output: dict[str, float] = field(default_factory=dict)

    def intermediate_by_industry(self) -> dict[str, float]:
        out = {code: 0.0 for code in self.industries}
        for (_, icode), v in self.intermediate.items():
            out[icode] += v
        return out

    def value_added_by_industry(self, code: str = "B1") -> dict[str, float]:
        return {i: self.value_added.get((code, i), 0.0) for i in self.industries}


def read_use_table(path: str, year: int) -> UseTable:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = wb[f"Use Table {year}"]
    rows = list(sheet.iter_rows(values_only=True))
    h1, h2 = rows[0], rows[1]

    industry_cols: dict[int, str] = {}
    for j, h in enumerate(h1):
        code = str(h or "").strip()
        if _INDUSTRY_CODE.match(code):
            industry_cols[j] = code
    trailing_cols: dict[int, str] = {}
    if industry_cols:
        for j in range(max(industry_cols) + 1, len(h2)):
            lab = str(h2[j] or "").strip()
            if lab:
                trailing_cols[j] = lab
    fd_cols = {j: lab for j, lab in trailing_cols.items()
               if not lab.lower().startswith("total")}

    table = UseTable(
        year=year,
        products=[],
        industries=list(industry_cols.values()),
        fd_categories=list(fd_cols.values()),
    )

    in_va_block = False
    for r in rows[3:]:
        code = str(r[0] or "").strip()
        desc = str(r[2] or "").strip()
        label2 = str(r[1] or "").strip()
        if _PRODUCT_CODE.match(code) and desc:
            table.products.append(Product(code=code, sic=label2, description=desc))
            if isinstance(r[3], (int, float)):
                table.total_supply_purchasers[code] = float(r[3])
            for j, icode in industry_cols.items():
                v = r[j]
                if isinstance(v, (int, float)) and v != 0:
                    table.intermediate[(code, icode)] = float(v)
            for j, lab in trailing_cols.items():
                v = r[j]
                if not isinstance(v, (int, float)):
                    continue
                if lab in fd_cols.values():
                    if v != 0:
                        table.final_demand[(code, lab)] = float(v)
                elif lab.lower().startswith("total industry"):
                    table.embedded_intermediate_by_product[code] = float(v)
            continue
        if code == "P2" and "total uses" in label2.lower():
            for j, icode in industry_cols.items():
                if isinstance(r[j], (int, float)):
                    table.embedded_intermediate_by_industry[icode] = float(r[j])
            in_va_block = True
            continue
        if code == "P1" and "total output" in label2.lower():
            for j, icode in industry_cols.items():
                if isinstance(r[j], (int, float)):
                    table.embedded_output[icode] = float(r[j])
            in_va_block = False
            continue
        if in_va_block and code:
            table.va_components[code] = label2
            for j, icode in industry_cols.items():
                v = r[j]
                if isinstance(v, (int, float)) and v != 0:
                    table.value_added[(code, icode)] = float(v)

    return table


def check_use_table(table: UseTable, tolerance: float = 0.5) -> list[Finding]:
    findings: list[Finding] = []
    inter_by_prod: dict[str, float] = {p.code: 0.0 for p in table.products}
    for (pcode, _), v in table.intermediate.items():
        inter_by_prod[pcode] += v
    fd_by_prod: dict[str, float] = {p.code: 0.0 for p in table.products}
    for (pcode, _), v in table.final_demand.items():
        fd_by_prod[pcode] += v
    for p in table.products:
        supply = table.total_supply_purchasers.get(p.code)
        if supply is None:
            findings.append(Finding("use-row-identity", p.code, "missing total supply cell"))
            continue
        gap = supply - (inter_by_prod[p.code] + fd_by_prod[p.code])
        if abs(gap) > tolerance:
            findings.append(Finding(
                "use-row-identity", p.code,
                f"total supply {supply:,.1f} != intermediate {inter_by_prod[p.code]:,.1f} "
                f"+ final demand {fd_by_prod[p.code]:,.1f} (gap {gap:,.1f})"))
    inter_by_ind = table.intermediate_by_industry()
    b1 = table.value_added_by_industry("B1")
    for icode in table.industries:
        emb = table.embedded_intermediate_by_industry.get(icode)
        if emb is not None and abs(inter_by_ind[icode] - emb) > tolerance:
            findings.append(Finding(
                "use-column-total", icode,
                f"recomputed intermediate {inter_by_ind[icode]:,.1f} != embedded {emb:,.1f}"))
        out = table.embedded_output.get(icode)
        if out is not None:
            gap = out - (inter_by_ind[icode] + b1[icode])
            if abs(gap) > tolerance:
                findings.append(Finding(
                    "output-identity", icode,
                    f"output {out:,.1f} != intermediate {inter_by_ind[icode]:,.1f} "
                    f"+ B1 {b1[icode]:,.1f} (gap {gap:,.1f})"))
        va_parts = sum(table.value_added.get((c, icode), 0.0) for c in ("D1", "D2/3", "B2/3"))
        if ("B1" in table.va_components
                and abs(va_parts - b1[icode]) > tolerance):
            findings.append(Finding(
                "va-identity", icode,
                f"D1+D2/3+B2/3 {va_parts:,.1f} != B1 {b1[icode]:,.1f}"))
    return findings


def cross_check_supply_use(supply: SupplyTable, use: UseTable,
                           tolerance: float = 0.5) -> list[Finding]:
    """Identities that tie the two tables of one year together."""
    findings: list[Finding] = []
    sup_out = supply.output_by_industry()
    for icode, out in use.embedded_output.items():
        gap = sup_out.get(icode, 0.0) - out
        if abs(gap) > tolerance:
            findings.append(Finding(
                "supply-use-output", icode,
                f"supply-table output {sup_out.get(icode, 0.0):,.1f} != "
                f"use-table output {out:,.1f}"))
    for p in use.products:
        s = supply.total_supply_purchasers.get(p.code)
        u = use.total_supply_purchasers.get(p.code)
        if s is not None and u is not None and abs(s - u) > tolerance:
            findings.append(Finding(
                "supply-use-product-total", p.code,
                f"supply table {s:,.1f} != use table {u:,.1f}"))
    return findings
