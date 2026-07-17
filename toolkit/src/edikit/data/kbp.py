"""Reader for SARB Quarterly Bulletin bulk data workbooks (KBP series).

The SARB distributes Quarterly Bulletin time series as zipped xlsx files per
section (e.g. "05KBP6 National Accounts December 2022.zip"). Each workbook
holds wide sheets: one row per period, one column per series code, with the
first column a date stamp (YYYYMMDD; annual sheets use one row per year).
Annual series (J-suffixed codes) live on sheets named like "J1".
"""

from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass, field

import openpyxl


@dataclass
class KbpData:
    """Annual KBP observations: values[code][year] -> value (R millions)."""

    values: dict[str, dict[int, float]] = field(default_factory=dict)
    sources: dict[str, str] = field(default_factory=dict)  # code -> file it came from

    def get(self, code: str, year: int) -> float:
        try:
            return self.values[code][year]
        except KeyError:
            raise KeyError(f"{code} has no observation for {year}") from None

    def codes(self) -> list[str]:
        return sorted(self.values)


def _read_annual_sheet(ws, source_name: str, data: KbpData) -> None:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return
    col_codes = {}
    for j, h in enumerate(header):
        code = str(h or "").strip()
        if j > 0 and code:
            col_codes[j] = code
    for r in rows:
        stamp = str(r[0] or "").strip()
        if len(stamp) < 4 or not stamp[:4].isdigit():
            continue
        year = int(stamp[:4])
        for j, code in col_codes.items():
            v = r[j]
            if isinstance(v, (int, float)):
                data.values.setdefault(code, {})[year] = float(v)
                data.sources.setdefault(code, source_name)


def read_kbp_workbook(fileobj_or_path, source_name: str = "",
                      data: KbpData | None = None) -> KbpData:
    """Read annual (J*) sheets of one KBP workbook into a KbpData."""
    data = data if data is not None else KbpData()
    wb = openpyxl.load_workbook(fileobj_or_path, data_only=True, read_only=True)
    for name in wb.sheetnames:
        if name.strip().upper().startswith("J"):
            _read_annual_sheet(wb[name], source_name, data)
    return data


def read_kbp_zips(zip_paths: list[str]) -> KbpData:
    """Read the annual sheets of every xlsx inside the given SARB zips."""
    data = KbpData()
    for zp in zip_paths:
        with zipfile.ZipFile(zp) as z:
            for name in z.namelist():
                if name.lower().endswith((".xlsx", ".xls")) and not name.startswith("__MACOSX"):
                    with z.open(name) as f:
                        read_kbp_workbook(io.BytesIO(f.read()), source_name=f"{zp}::{name}",
                                          data=data)
    return data


def evaluate_formula(formula: str, data: KbpData, year: int) -> float:
    """Evaluate a signed sum of KBP codes, e.g. '+KBP6001J+KBP6002J-KBP6601J'."""
    total = 0.0
    token = ""
    sign = 1
    for ch in formula.replace(" ", "") + "+":
        if ch in "+-":
            if token:
                total += sign * data.get(token, year)
            token = ""
            sign = 1 if ch == "+" else -1
        else:
            token += ch
    return total
