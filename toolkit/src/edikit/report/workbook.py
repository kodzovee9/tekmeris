"""Transparent Excel workbook generation.

Writes auditable workbooks per the project's reproducibility standard:
formulas stay live (totals and balance checks recompute in Excel, so an
auditor sees the checks pass rather than trusting pasted numbers), inputs
and checks live on separate sheets, and every workbook carries a README
sheet stating what it is and how to regenerate it.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BOLD = Font(bold=True)


def add_readme_sheet(wb: Workbook, title: str, lines: list[str]) -> None:
    ws = wb.active if wb.sheetnames == ["Sheet"] else wb.create_sheet()
    ws.title = "README"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110


def add_table_sheet(wb: Workbook, name: str, header: list[str],
                    rows: list[list], widths: list[int] | None = None) -> None:
    ws = wb.create_sheet(name)
    for j, h in enumerate(header, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = BOLD
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    for j, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"


def add_matrix_sheet(wb: Workbook, name: str, accounts: list[str],
                     cells: dict[tuple[str, str], float],
                     number_format: str = "#,##0.0") -> None:
    """Square labeled matrix with LIVE total formulas.

    Layout: row 1 = column labels, column A = row labels; a Total column and
    Total row computed with SUM formulas; a 'Row-Col gap' column with live
    balance-check formulas referencing the totals.
    """
    ws = wb.create_sheet(name)
    n = len(accounts)
    tot_col = n + 2                 # column index of "Total"
    gap_col = n + 3
    tot_row = n + 2
    ws.cell(row=1, column=1, value=f"{name} (R million)").font = BOLD
    for j, a in enumerate(accounts, start=2):
        ws.cell(row=1, column=j, value=a).font = BOLD
    ws.cell(row=1, column=tot_col, value="Total").font = BOLD
    ws.cell(row=1, column=gap_col, value="Row-Col gap").font = BOLD
    index = {a: i for i, a in enumerate(accounts)}
    for i, a in enumerate(accounts, start=2):
        ws.cell(row=i, column=1, value=a).font = BOLD
    for (r, c), v in cells.items():
        if r in index and c in index and v:
            cell = ws.cell(row=index[r] + 2, column=index[c] + 2, value=round(v, 3))
            cell.number_format = number_format
    last_data_col = get_column_letter(n + 1)
    for i in range(2, n + 2):
        f = ws.cell(row=i, column=tot_col,
                    value=f"=SUM(B{i}:{last_data_col}{i})")
        f.number_format = number_format
        colL = get_column_letter(i)  # symmetric position of this account's column
        g = ws.cell(row=i, column=gap_col,
                    value=f"={get_column_letter(tot_col)}{i}-{colL}{tot_row}")
        g.number_format = number_format
    for j in range(2, n + 2):
        colL = get_column_letter(j)
        f = ws.cell(row=tot_row, column=j, value=f"=SUM({colL}2:{colL}{n + 1})")
        f.number_format = number_format
    ws.cell(row=tot_row, column=1, value="Total").font = BOLD
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 10
