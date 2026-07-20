import openpyxl
import pytest

from edikit.data.concordance import Concordance
from edikit.data.sam import read_labeled_matrix
from edikit.data.sut import check_supply_table, read_supply_table
from edikit.model.accounts import AccountKind, classify_by_prefix


# ---------- concordance ----------

def test_concordance_aggregate_and_validate():
    c = Concordance(name="t", mapping={"I1": "a1", "I2": "a1", "I3": "a2"})
    assert c.aggregate({"I1": 1.0, "I2": 2.0, "I3": 4.0}) == {"a1": 3.0, "a2": 4.0}
    assert c.validate(["I1", "I2", "I3"]) == []
    problems = c.validate(["I1", "I2", "I3", "I4"])
    assert len(problems) == 1 and "I4" in problems[0]


def test_concordance_csv_roundtrip(tmp_path):
    c = Concordance(name="t", mapping={"I1": "a1"}, target_labels={"a1": "Agri"})
    p = tmp_path / "c.csv"
    c.to_csv(str(p))
    c2 = Concordance.from_csv(str(p))
    assert c2.mapping == c.mapping and c2.target_labels == c.target_labels


# ---------- accounts ----------

def test_classify_by_prefix_longest_wins_and_raises():
    rules = {"a": AccountKind.ACTIVITY, "at": AccountKind.TAX}
    s = classify_by_prefix(["agri", "atax"], rules)
    kinds = {a.code: a.kind for a in s.accounts}
    assert kinds == {"agri": AccountKind.ACTIVITY, "atax": AccountKind.TAX}
    with pytest.raises(ValueError):
        classify_by_prefix(["zzz"], rules)


# ---------- synthetic workbook fixtures ----------

def _make_sut_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supply Table 2019"
    ws.append(["", "", "", "Total supply at purchasers' prices",
               "Taxes less subsidies on products", "Trade and transport margins",
               "Total supply at basic prices", "I1", "I2"])
    ws.append(["Supply Table 2019"])
    ws.append(["R'million"])
    # basic = domestic output only in this toy (no imports)
    ws.append(["P1", "011", "Cereals", 130.0, 10.0, 20.0, 100.0, 60.0, 40.0])
    ws.append(["P2", "012", "Fruit", 62.0, 2.0, 10.0, 50.0, 50.0, 0.0])
    ws.append(["P1", "Total supply at basic prices", "", 192.0, 12.0, 30.0, 150.0, 110.0, 40.0])
    il = wb.create_sheet("Industry List")
    il.append(["Industry number in the SUT", "desc", "sic", "sic pub"])
    il.append(["I1", "Farming", "111", "11"])
    il.append(["I2", "Mining", "210", "21"])
    wb.save(path)


def test_read_and_check_supply_table(tmp_path):
    p = tmp_path / "sut.xlsx"
    _make_sut_workbook(str(p))
    t = read_supply_table(str(p), 2019)
    assert [pr.code for pr in t.products] == ["P1", "P2"]
    assert t.industries == ["I1", "I2"]
    assert t.output_by_industry() == {"I1": 110.0, "I2": 40.0}
    assert t.embedded_output_totals == {"I1": 110.0, "I2": 40.0}
    assert check_supply_table(t) == []


def test_check_supply_table_flags_broken_identity(tmp_path):
    p = tmp_path / "sut.xlsx"
    _make_sut_workbook(str(p))
    t = read_supply_table(str(p), 2019)
    t.total_supply_purchasers["P1"] = 999.0
    findings = check_supply_table(t)
    assert any(f.check == "row-identity" and f.subject == "P1" for f in findings)


def _make_sam_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAM"
    ws.append(["Contents"])
    ws.append([None, "aagri", "cceri", "total"])
    ws.append(["aagri", 0, 100.0, 100.0])
    ws.append(["cceri", 100.0, 0, 100.0])
    ws.append(["total", 100.0, 100.0])
    wb.save(path)


def test_read_labeled_matrix(tmp_path):
    p = tmp_path / "sam.xlsx"
    _make_sam_workbook(str(p))
    m = read_labeled_matrix(str(p), "SAM")
    assert m.row_accounts == ["aagri", "cceri"]
    assert m.col_accounts == ["aagri", "cceri"]
    assert m.cells == {("aagri", "cceri"): 100.0, ("cceri", "aagri"): 100.0}
    assert m.embedded_col_totals == {"aagri": 100.0, "cceri": 100.0}
    assert m.embedded_row_totals == {"aagri": 100.0, "cceri": 100.0}
    assert m.col_sums() == {"aagri": 100.0, "cceri": 100.0}
    assert m.accounts_with_prefix("a") == ["aagri"]


def _make_use_workbook(path):
    from edikit.data.sut import read_use_table  # noqa: F401
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Use Table 2019"
    ws.append(["", "", "", "Total supply at purchasers' prices",
               "Taxes less subsidies on products", "I1", "I2", "", "", ""])
    ws.append(["", "", "", "", "", "11", "21", "Total industry", "Exports", "Households"])
    ws.append(["R'million"])
    # P1: supply 130 = intermediate (40+30) + exports 40 + hh 20
    ws.append(["P1", "011", "Cereals", 130.0, 10.0, 40.0, 30.0, 70.0, 40.0, 20.0])
    # P2x: supply 62 = intermediate (10+2) + exports 0 + hh 50
    ws.append(["P2", "012", "Fruit", 62.0, 2.0, 10.0, 2.0, 12.0, 0.0, 50.0])
    ws.append(["P2", " Total uses at purchasers' prices", "", None, None, 50.0, 32.0])
    ws.append(["B1", " Gross value added at basic prices", "", None, None, 60.0, 8.0])
    ws.append(["D1", "    Compensation of employees", "", None, None, 35.0, 5.0])
    ws.append(["D2/3", "    Taxes less subsidies", "", None, None, 5.0, 1.0])
    ws.append(["B2/3", "    Gross operation surplus/mixed income", "", None, None, 20.0, 2.0])
    ws.append(["P1", " Total output at basic prices", "", None, None, 110.0, 40.0])
    wb.save(path)


def test_read_and_check_use_table(tmp_path):
    from edikit.data.sut import check_use_table, read_use_table
    p = tmp_path / "use.xlsx"
    _make_use_workbook(str(p))
    t = read_use_table(str(p), 2019)
    assert [pr.code for pr in t.products] == ["P1", "P2"]
    assert t.industries == ["I1", "I2"]
    assert t.fd_categories == ["Exports", "Households"]
    assert t.intermediate_by_industry() == {"I1": 50.0, "I2": 32.0}
    assert t.value_added_by_industry("B1") == {"I1": 60.0, "I2": 8.0}
    assert t.embedded_output == {"I1": 110.0, "I2": 40.0}
    assert t.embedded_intermediate_by_product == {"P1": 70.0, "P2": 12.0}
    assert check_use_table(t) == []


def test_check_use_table_flags_broken_output_identity(tmp_path):
    from edikit.data.sut import check_use_table, read_use_table
    p = tmp_path / "use.xlsx"
    _make_use_workbook(str(p))
    t = read_use_table(str(p), 2019)
    t.embedded_output["I1"] = 999.0
    assert any(f.check == "output-identity" and f.subject == "I1"
               for f in check_use_table(t))


def test_cross_check_supply_use(tmp_path):
    from edikit.data.sut import cross_check_supply_use, read_supply_table, read_use_table
    ps, pu = tmp_path / "sut.xlsx", tmp_path / "use.xlsx"
    _make_sut_workbook(str(ps))
    _make_use_workbook(str(pu))
    sup = read_supply_table(str(ps), 2019)
    use = read_use_table(str(pu), 2019)
    assert cross_check_supply_use(sup, use) == []
    use.embedded_output["I2"] = 999.0
    assert any(f.check == "supply-use-output" for f in cross_check_supply_use(sup, use))


def _make_kbp_workbook(path):
    wb = openpyxl.Workbook()
    k = wb.active
    k.title = "K1"  # quarterly sheet, must be ignored by the annual reader
    k.append(["", "KBP9999K"])
    k.append(["20190100", 1.0])
    j = wb.create_sheet("J1")
    j.append(["", "KBP6000J", "KBP6001J"])
    j.append(["20180100", 2500000.0, 1000000.0])
    j.append(["20190100", 2732292.0, 1100000.0])
    wb.save(path)


def test_kbp_reader_and_formula(tmp_path):
    import zipfile
    from edikit.data.kbp import evaluate_formula, read_kbp_workbook, read_kbp_zips
    p = tmp_path / "kbp.xlsx"
    _make_kbp_workbook(str(p))
    d = read_kbp_workbook(str(p))
    assert d.get("KBP6000J", 2019) == 2732292.0
    assert d.get("KBP6000J", 2018) == 2500000.0
    assert "KBP9999K" not in d.values  # quarterly sheet ignored
    assert evaluate_formula("+KBP6000J+KBP6001J", d, 2019) == 3832292.0
    assert evaluate_formula("+KBP6000J-KBP6001J", d, 2019) == 1632292.0
    zp = tmp_path / "kbp.zip"
    with zipfile.ZipFile(zp, "w") as z:
        z.write(p, "inner/kbp.xlsx")
    d2 = read_kbp_zips([str(zp)])
    assert d2.get("KBP6001J", 2019) == 1100000.0
    with pytest.raises(KeyError):
        d2.get("KBP6001J", 1900)


# ---------- balancing ----------

def test_ras_converges_and_hits_targets():
    from edikit.model.balancing import ras
    seed = {("r1", "c1"): 1.0, ("r1", "c2"): 1.0, ("r2", "c1"): 1.0, ("r2", "c2"): 1.0}
    res = ras(seed, {"r1": 6.0, "r2": 4.0}, {"c1": 7.0, "c2": 3.0})
    assert res.converged
    assert abs(res.matrix[("r1", "c1")] + res.matrix[("r1", "c2")] - 6.0) < 1e-6
    assert abs(res.matrix[("r1", "c1")] + res.matrix[("r2", "c1")] - 7.0) < 1e-6
    assert res.max_factor_deviation() > 0


def test_ras_preserves_zero_structure():
    from edikit.model.balancing import ras
    seed = {("r1", "c1"): 2.0, ("r2", "c2"): 3.0}  # block-diagonal
    res = ras(seed, {"r1": 5.0, "r2": 7.0}, {"c1": 5.0, "c2": 7.0})
    assert res.converged
    assert ("r1", "c2") not in res.matrix and ("r2", "c1") not in res.matrix
    assert abs(res.matrix[("r1", "c1")] - 5.0) < 1e-6


def test_ras_rejects_bad_inputs():
    from edikit.model.balancing import ras
    with pytest.raises(ValueError):
        ras({("r", "c"): -1.0}, {"r": 1.0}, {"c": 1.0})
    with pytest.raises(ValueError):
        ras({("r", "c"): 1.0}, {"r": 2.0}, {"c": 1.0})  # inconsistent totals
    with pytest.raises(ValueError):
        ras({("r", "c"): 1.0}, {"r": 1.0, "r2": 1.0}, {"c": 2.0})  # empty seed row


def test_gras_reduces_to_ras_on_nonnegative_seed():
    from edikit.model.balancing import gras, ras
    seed = {("r1", "c1"): 1.0, ("r1", "c2"): 1.0, ("r2", "c1"): 1.0, ("r2", "c2"): 1.0}
    rt, ct = {"r1": 6.0, "r2": 4.0}, {"c1": 7.0, "c2": 3.0}
    a, b = ras(seed, rt, ct), gras(seed, rt, ct)
    assert a.converged and b.converged
    assert all(abs(a.matrix[k] - b.matrix[k]) < 1e-6 for k in a.matrix)


def test_gras_handles_negative_seed_and_hits_margins():
    from edikit.model.balancing import gras
    # seed with a genuine negative entry (as a real SAM has: margins, cif/fob)
    seed = {("r1", "c1"): 10.0, ("r1", "c2"): -2.0,
            ("r2", "c1"): 3.0, ("r2", "c2"): 5.0}
    rt, ct = {"r1": 12.0, "r2": 16.0}, {"c1": 20.0, "c2": 8.0}
    g = gras(seed, rt, ct)
    assert g.converged
    rs, cs = {}, {}
    for (r, c), v in g.matrix.items():
        rs[r] = rs.get(r, 0.0) + v
        cs[c] = cs.get(c, 0.0) + v
    assert all(abs(rs[r] - rt[r]) < 1e-6 for r in rt)
    assert all(abs(cs[c] - ct[c]) < 1e-6 for c in ct)
    assert g.matrix[("r1", "c2")] < 0        # sign of the seed cell is preserved


def test_gras_rejects_inconsistent_totals():
    from edikit.model.balancing import gras
    with pytest.raises(ValueError):
        gras({("r", "c"): 1.0}, {"r": 2.0}, {"c": 1.0})


def test_workbook_writer_live_formulas(tmp_path):
    from openpyxl import Workbook, load_workbook
    from edikit.report.workbook import add_matrix_sheet, add_readme_sheet, add_table_sheet
    wb = Workbook()
    add_readme_sheet(wb, "Test book", ["line one", "line two"])
    add_table_sheet(wb, "T", ["a", "b"], [[1, 2], [3, 4]])
    add_matrix_sheet(wb, "M", ["x", "y"], {("x", "y"): 5.0, ("y", "x"): 5.0})
    p = tmp_path / "t.xlsx"
    wb.save(p)
    rd = load_workbook(p)
    assert rd.sheetnames == ["README", "T", "M"]
    ws = rd["M"]
    assert ws["D2"].value == "=SUM(B2:C2)"      # row total formula, live
    assert ws["B4"].value == "=SUM(B2:B3)"      # column total formula, live
    assert ws["E2"].value == "=D2-B4"           # balance check formula, live
    assert ws["C2"].value == 5.0


def test_eurostat_jsonstat_reader(tmp_path):
    import json
    from edikit.data.eurostat import read_jsonstat
    doc = {
        "label": "Toy table", "id": ["geo", "item"], "size": [1, 3],
        "dimension": {
            "geo": {"category": {"index": {"NL": 0}, "label": {"NL": "Netherlands"}}},
            "item": {"category": {"index": {"A": 0, "B": 1, "C": 2},
                                  "label": {"A": "Alpha", "B": "Beta", "C": "Gamma"}}},
        },
        "value": {"0": 1.5, "2": 3.0},
    }
    p = tmp_path / "t.json"
    p.write_text(json.dumps(doc))
    t = read_jsonstat(str(p))
    assert t.values == {("NL", "A"): 1.5, ("NL", "C"): 3.0}
    assert t.value(geo="NL", item="A") == 1.5
    assert t.value(geo="NL", item="B") is None
    assert t.slice(geo="NL") == {("A",): 1.5, ("C",): 3.0}
    assert t.labels["item"]["B"] == "Beta"


def test_nace_coverage_parsing():
    from edikit.pipeline.eurostat_sam import coverage
    assert coverage("D") == ("D", None)
    assert coverage("CPA_C10-12") == ("C", {10, 11, 12})
    assert coverage("C31_32") == ("C", {31, 32})
    assert coverage("J59_60") == ("J", {59, 60})
    assert coverage("L68A") == ("L", {68})
    assert coverage("CPA_A01") == ("A", {1})


def test_partition_drops_contained_codes():
    from edikit.pipeline.eurostat_sam import partition
    # sub-details contained in a present aggregate are dropped
    assert partition(["C10-12", "C11", "C13-15", "C16"]) == \
        ["C10-12", "C13-15", "C16"]
    # a whole-section code absorbs its divisions
    assert partition(["E", "E36", "E37-39"]) == ["E"]
    # disjoint codes all survive, across sections
    assert partition(["A01", "A02", "B", "C10-12"]) == \
        ["A01", "A02", "B", "C10-12"]


def test_audit_balance_grain_and_aggregates(tmp_path):
    from edikit.pipeline.audit import audit, read_long_csv, write_report
    p = tmp_path / "sam.csv"
    # tiny 2-sector economy: act sells 100 to com; com sells 60 to hhd,
    # 40 back to act as intermediates; factors 60 -> hhd; hhd saves 0
    p.write_text("row,col,value\n"
                 "act,com,100\ncom,act,40\ncom,hhd,60\n"
                 "fac,act,60\nhhd,fac,60\n")
    cells = read_long_csv(p)
    kinds = {"act": "activity", "com": "commodity",
             "fac": "factor", "hhd": "household"}
    res = audit(cells, kinds=kinds)
    assert res.grain == 10.0
    assert res.gaps["act"] == 0 and res.gaps["com"] == 0
    assert res.aggregates["GVA (factor payments by activities)"] == 60
    assert res.aggregates["Private consumption"] == 60
    assert not res.coverage_gap
    write_report(res, tmp_path / "audit.md")
    text = (tmp_path / "audit.md").read_text()
    assert "Accounts: 4" in text and "GVA" in text


def test_audit_flags_unclassified_accounts():
    from edikit.pipeline.audit import audit
    res = audit({("a", "b"): 1.0, ("b", "a"): 1.0}, kinds={"a": "activity"})
    assert res.coverage_gap == ["b"]


def test_macro_balance_flips_negatives_and_converges():
    from edikit.pipeline.eurostat_sam import MacroSAMResult, balance
    res = MacroSAMResult(
        country="XX", year=2019, inds=[], prods=[], findings=[],
        cross_diffs=[], n3_check=0.0, n_closed=0, prod_resid={},
        sam={("a", "b"): 100.0, ("b", "a"): 90.0, ("a", "c"): -10.0,
             ("c", "b"): 12.0, ("b", "c"): 3.0})
    bal, flipped = balance(res)
    assert flipped == [("a", "c")]
    assert ("c", "a") in bal.matrix and all(v >= 0 for v in bal.matrix.values())
    assert bal.converged and max(bal.max_row_gap, bal.max_col_gap) < 1e-6


def _jsonstat(dims: dict, values: dict) -> dict:
    """Build a minimal JSON-stat 2.0 document from {(coords...): value}."""
    ids = list(dims)
    sizes = [len(dims[d]) for d in ids]
    doc = {"label": "synthetic", "id": ids, "size": sizes,
           "dimension": {d: {"category": {"index": {c: i for i, c in
                                                    enumerate(dims[d])}}}
                         for d in ids},
           "value": {}}
    for key, v in values.items():
        lin = 0
        for d, c in zip(ids, key):
            lin = lin * len(dims[d]) + dims[d].index(c)
        doc["value"][str(lin)] = v
    return doc


def test_eurostat_generate_and_balance_on_synthetic_economy(tmp_path):
    """Two-industry economy whose identities close exactly: generate()
    must find zero findings, full closure, GDP = published B1G, and
    balance() must converge with the accounts already near-balanced."""
    import json
    from edikit.pipeline.eurostat_sam import balance, generate

    fixed = {"freq": ["A"], "unit": ["MIO_EUR"], "stk_flow": ["TOTAL"]}
    geo_time = {"geo": ["XX"], "time": ["2019"]}
    F = ("A", "MIO_EUR", "TOTAL")
    GT = ("XX", "2019")

    sup_vals = {}
    for ind, prd, v in [
            ("A01", "CPA_A01", 100), ("B", "CPA_B", 200),
            ("TOTAL", "CPA_A01", 100), ("TOTAL", "CPA_B", 200),
            ("TOTAL", "CPA_TOTAL", 300),
            ("P7", "CPA_A01", 10), ("P7", "CPA_B", 20),
            ("P7", "CPA_TOTAL", 30),
            ("D21X31", "CPA_A01", 5), ("D21X31", "CPA_B", 10),
            ("D21X31", "CPA_TOTAL", 15),
            ("TS_BP", "CPA_A01", 110), ("TS_BP", "CPA_B", 220),
            ("TS_BP", "CPA_TOTAL", 330),
            ("TS_PP", "CPA_A01", 115), ("TS_PP", "CPA_B", 230),
            ("TS_PP", "CPA_TOTAL", 345)]:
        sup_vals[F + (ind, prd) + GT] = v
    sup = _jsonstat({**fixed, "ind_impv": ["TOTAL", "TS_BP", "TS_PP",
                                           "P7", "D21X31", "A01", "B"],
                     "prd_amo": ["CPA_TOTAL", "CPA_A01", "CPA_B"],
                     **geo_time}, sup_vals)

    use_vals = {}
    for ind, prd, v in [
            # intermediates and their published TOTAL column
            ("A01", "CPA_B", 30), ("B", "CPA_A01", 40),
            ("TOTAL", "CPA_A01", 40), ("TOTAL", "CPA_B", 30),
            ("TOTAL", "CPA_TOTAL", 70),
            # value added: B1G = D1 + D29X39 + B2A3G; P1 = inter + B1G
            ("A01", "D1", 50), ("A01", "D29X39", 5), ("A01", "B2A3G", 15),
            ("A01", "B1G", 70), ("A01", "P1", 100),
            ("B", "D1", 100), ("B", "D29X39", 10), ("B", "B2A3G", 50),
            ("B", "B1G", 160), ("B", "P1", 200),
            ("TOTAL", "D1", 150), ("TOTAL", "D29X39", 15),
            ("TOTAL", "B2A3G", 65), ("TOTAL", "B1G", 230),
            ("TOTAL", "P1", 300),
            # final demand: TS_PP = intermediates + final demand
            ("P3_S14", "CPA_A01", 50), ("P3_S14", "CPA_B", 120),
            ("P3_S13", "CPA_A01", 10), ("P3_S13", "CPA_B", 20),
            ("P51G", "CPA_A01", 10), ("P51G", "CPA_B", 30),
            ("P6", "CPA_A01", 5), ("P6", "CPA_B", 30),
            ("P3_S14", "CPA_TOTAL", 170), ("P3_S13", "CPA_TOTAL", 30),
            ("P51G", "CPA_TOTAL", 40), ("P6", "CPA_TOTAL", 35)]:
        use_vals[F + (ind, prd) + GT] = v
    use = _jsonstat({**fixed, "ind_use": ["TOTAL", "P3_S13", "P3_S14",
                                          "P3_S15", "P51G", "P52", "P53",
                                          "P6", "A01", "B"],
                     "prd_use": ["CPA_TOTAL", "CPA_A01", "CPA_B", "D1",
                                 "D29X39", "B2A3G", "B1G", "P1"],
                     **geo_time}, use_vals)

    nasa_vals = {}
    for item, sector, direct, v in [
            ("D1", "S14_S15", "RECV", 152), ("D1", "S2", "PAID", 2),
            ("B2A3G", "S11", "RECV", 40), ("B2A3G", "S12", "RECV", 5),
            ("B2A3G", "S13", "RECV", 5), ("B2A3G", "S14_S15", "RECV", 15),
            ("D4", "S11", "PAID", 10), ("D4", "S14_S15", "RECV", 8),
            ("D4", "S2", "RECV", 2),
            ("B8G", "S11", "RECV", 30), ("B8G", "S12", "RECV", 5),
            ("B8G", "S13", "RECV", 5), ("B8G", "S14_S15", "RECV", 5),
            ("B9", "S1", "PAID", 5)]:
        nasa_vals[("A", "CP_MEUR", direct, item, sector) + GT] = v
    nasa = _jsonstat({"freq": ["A"], "unit": ["CP_MEUR"],
                      "direct": ["PAID", "RECV"],
                      "na_item": ["D1", "B2A3G", "D4", "B8G", "B9"],
                      "sector": ["S1", "S11", "S12", "S13", "S14_S15", "S2"],
                      **geo_time}, nasa_vals)

    paths = {}
    for name, doc in [("supply", sup), ("use", use), ("sectors", nasa)]:
        p = tmp_path / f"{name}.json"
        p.write_text(json.dumps(doc))
        paths[name] = p

    res = generate("XX", 2019, paths)
    assert res.inds == ["A01", "B"] and len(res.prods) == 2
    assert res.findings == [] and res.cross_diffs == []
    assert res.n3_check == 0 and res.n_closed == 2
    assert res.gdp == 230          # = published B1G total
    assert not res.coverage        # nothing suppressed
    # the fixture is engineered so EVERY account balances exactly
    assert max(abs(v) for v in res.imbalances.values()) < 1e-9
    bal, flipped = balance(res)
    assert bal.converged and not flipped
    assert max(bal.max_row_gap, bal.max_col_gap) < 1e-6
    assert bal.max_factor_deviation() < 1e-9   # nothing to adjust
    # the writers: report with balancing section, balanced CSV with factors
    from edikit.pipeline.eurostat_sam import write_balanced_csv, write_report
    write_report(res, tmp_path / "r.md", balance=bal, flipped=flipped)
    text = (tmp_path / "r.md").read_text()
    assert "Balanced macro SAM" in text and "converged=True" in text
    write_balanced_csv(bal, tmp_path / "b.csv")
    lines = (tmp_path / "b.csv").read_text().splitlines()
    assert lines[0] == "row,col,value_EURm,ras_factor"
    assert len(lines) == len(bal.matrix) + 1


def test_audit_matrix_reader_kinds_and_controls(tmp_path):
    from edikit.pipeline.audit import (audit, read_kinds, read_matrix_csv,
                                       write_report)
    (tmp_path / "sam.csv").write_text(
        "SAM,Code,act,com,Total\n"
        "Activities,act,,100,100\n"
        "Commodities,com,100,,100\n"
        "Total,Total,100,100,\n")
    cells = read_matrix_csv(tmp_path / "sam.csv", code_col=True)
    assert cells == {("act", "com"): 100.0, ("com", "act"): 100.0}
    (tmp_path / "kinds.csv").write_text(
        "account,kind\nact,activity\ncom,commodity\n")
    kinds = read_kinds(tmp_path / "kinds.csv")
    res = audit(cells, kinds=kinds,
                controls={"GVA (factor payments by activities)": 50.0})
    assert res.grain == 100.0 and res.max_gap[1] == 0
    write_report(res, tmp_path / "a.md", title="T")
    text = (tmp_path / "a.md").read_text()
    assert "Control" in text and "50.0" in text


def test_ansd_tre_reader_locates_blocks_and_checks_identities(tmp_path):
    """Synthetic 2-product x 2-branch TRE in the ANSD layout: the reader must
    locate the supply/use/VA blocks by label and the identities must close."""
    import openpyxl
    from edikit.data.ansd import read_tre, check_tre
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2014"

    def put(r, c, v):
        ws.cell(row=r, column=c, value=v)

    # branch descriptions (row 5) and branch codes (row 6, cols 12-13)
    put(5, 12, "BRANCHE A"); put(5, 13, "BRANCHE B")
    put(6, 12, "A00"); put(6, 13, "B00")
    # --- supply block: product rows 7-8 ---
    # cols: 3 purchaser, 11 basic, 12-13 make, 14 output, 15 cif, 16 imports
    for r, code, desc, pur, basic, mkA, mkB, out, imp in [
            (7, "A00", "Prod A", 110, 110, 100, 0, 100, 10),
            (8, "B00", "Prod B", 220, 220, 0, 200, 200, 20)]:
        put(r, 1, code); put(r, 2, desc); put(r, 3, pur); put(r, 11, basic)
        put(r, 12, mkA); put(r, 13, mkB); put(r, 14, out); put(r, 15, 0); put(r, 16, imp)
    # --- use block ---
    put(10, 1, "Emploi des produits")
    put(10, 16, "Exportations"); put(10, 17, "Consommation finale")
    put(10, 18, "Menages Sous-total"); put(10, 19, "Autoconsommation")
    put(10, 20, "Commercialisee"); put(10, 21, "Administrations")
    put(10, 22, "ISBL"); put(10, 23, "Formation brute de capital fixe")
    put(10, 24, "Variations des stocks"); put(10, 25, "objets de valeur")
    for r, code, pur, ucA, ucB, hh in [
            (12, "A00", 110, 30, 40, 40), (13, "B00", 220, 50, 60, 110)]:
        put(r, 1, code); put(r, 3, pur); put(r, 12, ucA); put(r, 13, ucB); put(r, 18, hh)
    # --- value-added block (cols 12-13 = branches) ---
    for r, label, a, b in [
            (15, "Valeur ajoutee brute /PIB", 20, 100),
            (16, "Remuneration des salaries", 10, 40),
            (17, "Salaires bruts", 10, 40),
            (18, "Contributions sociales effectives", 0, 0),
            (19, "Contributions sociales imputees", 0, 0),
            (20, "Impots sur la production", 2, 10),
            (21, "Subventions sur la production", 0, 0),
            (22, "Excedent brut d'exploitation / revenu mixte", 8, 50)]:
        put(r, 1, label); put(r, 12, a); put(r, 13, b)

    p = tmp_path / "tre.xlsx"
    wb.save(p)

    tre = read_tre(str(p), 2014)
    assert tre.products == ["A00", "B00"]
    assert tre.branches == ["A00", "B00"]
    assert tre.make[("A00", "A00")] == 100 and tre.make[("B00", "B00")] == 200
    assert tre.use[("B00", "A00")] == 50
    assert tre.value_added == {"A00": 20.0, "B00": 100.0}
    assert tre.operating_surplus == {"A00": 8.0, "B00": 50.0}
    assert tre.cons_households["B00"] == 110
    assert tre.output_by_branch() == {"A00": 100.0, "B00": 200.0}
    assert check_tre(tre) == []          # all four identities close

    # break the VA decomposition and confirm the check catches it
    tre.operating_surplus["A00"] = 999.0
    findings = check_tre(tre)
    assert any(f.check == "va-decomposition" and f.subject == "A00" for f in findings)


def test_ansd_tcei_reader_tracks_account_and_direction(tmp_path):
    """Synthetic institutional-sector sheet: the reader must key flows by
    (sector, account, direction, code), keeping the same code distinct across
    accounts and resources/emplois sides."""
    import openpyxl
    from edikit.data.ansd import read_tcei
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "0S1004"                       # households -> Men

    def put(r, c, v):
        ws.cell(row=r, column=c, value=v)

    put(4, 2, "Code"); put(4, 3, "Operations"); put(4, 4, 2014); put(4, 5, 2015)
    put(5, 3, "Compte d'affectation des revenus primaires")
    put(6, 3, "Ressources")
    put(7, 2, "D.1"); put(7, 3, "Remuneration"); put(7, 4, 100); put(7, 5, 110)
    put(8, 2, "D.4"); put(8, 3, "Revenus de la propriete"); put(8, 4, 50); put(8, 5, 55)
    put(9, 3, "Compte de distribution secondaire du revenu")
    put(10, 3, "Emplois")
    put(11, 2, "D.5"); put(11, 3, "Impots"); put(11, 4, 30); put(11, 5, 33)
    p = tmp_path / "tcei.xlsx"
    wb.save(p)

    t = read_tcei(str(p), 2014)
    assert t.get("Men", "primary", "resources", "D.1") == 100
    assert t.get("Men", "primary", "resources", "D.4") == 50
    assert t.get("Men", "secondary", "emplois", "D.5") == 30
    # a code that is not present, or the wrong direction, returns the default
    assert t.get("Men", "primary", "emplois", "D.1") == 0.0
    # the other year column is read independently
    assert read_tcei(str(p), 2015).get("Men", "primary", "resources", "D.1") == 110


def test_ansd_tcei_reader_handles_npish_colb_headers_and_row_sheet(tmp_path):
    """Two awkward sheet layouts: the NPISH sheet (0S1005) carries its account
    headers in column B rather than column C, and the rest-of-the-world sheet
    (0S2) uses the external-account block names. The reader must parse both."""
    import openpyxl
    from edikit.data.ansd import read_tcei
    wb = openpyxl.Workbook()

    # --- NPISH sheet: account header in column B ---
    ws = wb.active
    ws.title = "0S1005"

    def put(ws, r, c, v):
        ws.cell(row=r, column=c, value=v)

    put(ws, 4, 2, "Code"); put(ws, 4, 3, "Operations"); put(ws, 4, 4, 2014)
    put(ws, 5, 2, "Compte d'affectation des revenus primaires")   # header in col B
    put(ws, 6, 3, "Ressources")
    put(ws, 7, 2, "D.4"); put(ws, 7, 3, "Revenus de la propriete"); put(ws, 7, 4, 88)
    put(ws, 8, 2, "Compte de capital")                            # header in col B
    put(ws, 9, 3, "Ressources")
    put(ws, 10, 2, "B.8"); put(ws, 10, 3, "Epargne brute"); put(ws, 10, 4, 45)

    # --- rest-of-the-world sheet: external-account block names ---
    ws2 = wb.create_sheet("0S2")
    put(ws2, 4, 2, "Code"); put(ws2, 4, 3, "Operations"); put(ws2, 4, 4, 2014)
    put(ws2, 5, 3, "Compte exterieur des revenus primaires et des transferts")
    put(ws2, 6, 3, "Ressources")
    put(ws2, 7, 2, "D.1"); put(ws2, 7, 3, "Remuneration"); put(ws2, 7, 4, 19)
    put(ws2, 8, 3, "Emplois")
    put(ws2, 9, 2, "D.1"); put(ws2, 9, 3, "Remuneration"); put(ws2, 9, 4, 64)
    put(ws2, 10, 2, "B.12"); put(ws2, 10, 3, "Solde courant"); put(ws2, 10, 4, 670)

    p = tmp_path / "tcei_edge.xlsx"
    wb.save(p)
    t = read_tcei(str(p), 2014)
    # NPISH: header found in column B, capital-account saving read there
    assert t.get("ISBL", "primary", "resources", "D.4") == 88
    assert t.get("ISBL", "capital", "resources", "B.8") == 45
    # RoW: the one external-current block keeps the two D.1 directions distinct
    assert t.get("Rdm", "extcurrent", "resources", "D.1") == 19
    assert t.get("Rdm", "extcurrent", "emplois", "D.1") == 64
    assert t.get("Rdm", "extcurrent", "emplois", "B.12") == 670


# ---------- cross-producer reconciliation ----------

def test_nexus_classmap_maps_template_vocabulary():
    from edikit.pipeline.reconcile import nexus_classmap
    cases = {
        "amaiz": "ACT", "acons": "ACT", "cmaiz": "COM", "ctob": "COM",
        "flab-n": "LAB", "flab-s": "LAB", "fcap": "CAP", "flnd": "CAP",
        "hhd-r1": "HH", "hhd-u5": "HH", "ent": "ENT", "gov": "GOV",
        "row": "ROW", "s-i": "SAV", "dstk": "SAV",
        "dtax": "GOV", "mtax": "GOV", "stax": "GOV", "trc": "COM",
    }
    for code, kind in cases.items():
        assert nexus_classmap(code) == kind, code
    assert nexus_classmap("wat") is None       # outside the template


def test_macro_reduce_aggregates_and_flags_unclassified():
    from edikit.pipeline.reconcile import macro_reduce, nexus_classmap
    cells = {("amaiz", "cmaiz"): 10.0, ("arice", "cmaiz"): 5.0,  # both -> (ACT,COM)
             ("cmaiz", "hhd-r1"): 7.0, ("cmaiz", "hhd-u2"): 3.0,  # both -> (COM,HH)
             ("cmaiz", "zzz"): 99.0}                              # unknown col dropped
    m, unclassified = macro_reduce(cells, nexus_classmap)
    assert m[("ACT", "COM")] == 15.0
    assert m[("COM", "HH")] == 10.0
    assert unclassified == {"zzz"}
    assert ("COM", None) not in m and 99.0 not in m.values()


def test_national_accounts_counts_home_consumption_and_balances():
    from edikit.pipeline.reconcile import national_accounts
    # a balanced macro economy: activities make 100 of commodities; commodities
    # sell 40 to activities (intermediates), 55 to households, 10 to capital,
    # 15 as exports, minus 20 imports; households also buy 5 direct from
    # activities (home consumption); factors 60 -> households; product tax 5.
    m = {("ACT", "COM"): 100.0, ("COM", "ACT"): 40.0,
         ("LAB", "ACT"): 35.0, ("CAP", "ACT"): 25.0, ("GOV", "COM"): 5.0,
         ("COM", "HH"): 55.0, ("ACT", "HH"): 5.0, ("COM", "GOV"): 0.0,
         ("COM", "SAV"): 10.0, ("COM", "ROW"): 15.0, ("ROW", "COM"): 20.0}
    na = national_accounts(m)
    assert na["Value added (factor cost)"] == 60.0
    assert na["Net taxes on products"] == 5.0
    assert na["Household consumption"] == 60.0        # 55 commodity + 5 home
    assert na["GDP (income side)"] == 65.0
    # 60 hh + 0 gov + 10 capital + 15 exports - 20 imports = 65
    assert na["GDP (expenditure side)"] == 65.0


def test_scale_to_billions_reads_units():
    from edikit.pipeline.reconcile import scale_to_billions
    assert scale_to_billions("CFA MLN") == (1e-3, "millions")
    assert scale_to_billions("Billions of West African Cfa Franc")[0] == 1.0
    assert scale_to_billions("millions of dollars")[0] == 1e-3
    assert scale_to_billions("widgets") == (1.0, "widgets")   # unknown -> no rescale


def test_jrc_classmap_maps_template_vocabulary():
    from edikit.pipeline.reconcile import jrc_classmap
    cases = {
        "a_admn": "ACT", "ahf_DAK": "ACT",              # incl. household-farm activities
        "c_arch": "COM", "chrice": "COM", "trcost": "COM",
        "fl_SkillDAK": "LAB", "fl_UnSklZIG": "LAB", "flb_RWsk": "LAB",
        "fcp_ag": "CAP", "fl_irr": "CAP", "flivst": "CAP",
        "hh_RuralDAK": "HH", "hh_URDAKQ01": "HH",
        "ent": "ENT", "govt": "GOV", "dirtax": "GOV", "imptax": "GOV",
        "i_s": "SAV", "inv_road": "SAV", "row": "ROW", "hrow": "ROW",
    }
    for code, kind in cases.items():
        assert jrc_classmap(code) == kind, code
    assert jrc_classmap("zzz") is None


def test_read_jrc_sam_long_form(tmp_path):
    from edikit.pipeline.reconcile import read_jrc_sam
    p = tmp_path / "jrc.csv"
    # receiving = row (income), spending = col (expenditure); unit in header
    p.write_text(
        "Year,Spending Agent,Spending Agent (Code),Receiving Agent,"
        "Receiving Agent (Code),Value (CFA MLN)\n"
        "2014,Cereals comm,c_cer,Cereals act,a_cer,500\n"
        "2014,Cereals act,a_cer,Labour,fl_SkillDAK,120\n"
        "2014,Cereals act,a_cer,Cereals comm,c_cer,0\n")   # zero dropped
    cells, unit = read_jrc_sam(str(p))
    assert unit == "CFA MLN"
    assert cells == {("a_cer", "c_cer"): 500.0, ("fl_SkillDAK", "a_cer"): 120.0}


def test_reconcile_on_real_nexus_kenya_sam():
    """End-to-end on a genuine IFPRI Nexus SAM committed in the repo: the
    reduction must classify every account and the two GDP measures must agree
    to the file's rounding grain (values are integer billions)."""
    import pathlib
    p = (pathlib.Path(__file__).parents[2] / "papers" / "P001-sut-to-sam" / "data"
         / "raw" / "kenya" / "ifpri" / "002_IFPRI_SAM_KEN_2019_SAM.csv")
    if not p.exists():
        pytest.skip("Kenya Nexus SAM not present in this checkout")
    from edikit.pipeline.audit import read_matrix_csv
    from edikit.pipeline.reconcile import (macro_reduce, national_accounts,
                                           nexus_classmap)
    cells = read_matrix_csv(str(p), code_col=True)
    m, unclassified = macro_reduce(cells, nexus_classmap)
    assert not unclassified                       # whole template recognised
    na = national_accounts(m)
    gi, ge = na["GDP (income side)"], na["GDP (expenditure side)"]
    assert gi > 9000 and abs(gi - ge) / gi < 2e-3  # agree within rounding
    assert na["Imports"] > na["Exports"] > 0       # Kenya runs a trade deficit


def test_ansd_mcs_reader_reads_square_matrix(tmp_path):
    """Synthetic MCS: row codes in column B, the square matrix in the columns
    that follow in the same account order, then a TOTAL row/column."""
    import openpyxl
    from edikit.data.ansd import read_mcs
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2014"
    accounts = ["J1", "I1", "Men", "TOTAL"]
    # labels in column B (index 2 in openpyxl 1-based), matrix from column C
    for i, a in enumerate(accounts):
        ws.cell(row=8 + i, column=2, value=a)
    # J1 <- I1 = 100 ; I1 <- J1 = 40 ; I1 <- Men = 60 ; Men <- (nothing)
    ws.cell(row=8, column=4, value=100)      # J1 row, I1 col (col C+1)
    ws.cell(row=9, column=3, value=40)       # I1 row, J1 col
    ws.cell(row=9, column=5, value=60)       # I1 row, Men col
    p = tmp_path / "mcs.xlsx"
    wb.save(p)

    m = read_mcs(str(p), 2014)
    assert m.accounts == ["J1", "I1", "Men"]      # TOTAL excluded
    assert m.cells[("J1", "I1")] == 100
    assert m.cells[("I1", "J1")] == 40
    assert m.cells[("I1", "Men")] == 60
    assert ("Men", "J1") not in m.cells


def test_write_reconciliation_emits_both_producers_and_divergence(tmp_path):
    """The reconciliation report writer: two producers on the common grid
    should yield a report naming both, carrying the national aggregates, and
    surfacing unclassified accounts, mapping rules and notes.

    NB: ``year`` is accepted but never written to the report -- callers put
    the year in ``title``. Asserted absent here so the dead parameter is
    recorded rather than silently tolerated."""
    from edikit.pipeline.reconcile import Producer, write_reconciliation
    m = {("ACT", "COM"): 100.0, ("COM", "ACT"): 40.0,
         ("LAB", "ACT"): 35.0, ("CAP", "ACT"): 25.0,
         ("COM", "HH"): 55.0, ("COM", "SAV"): 10.0,
         ("COM", "ROW"): 15.0, ("ROW", "COM"): 20.0}
    a = Producer(name="Office", macro_sam=m, unit="bn", scale=1.0)
    b = Producer(name="Independent", macro_sam={k: v * 1.1 for k, v in m.items()},
                 unit="bn", scale=1.0, unclassified={"zzz"})
    out = tmp_path / "rec.md"
    write_reconciliation(out, [a, b], title="Test reconciliation", year=2018,
                         unit="bn", notes=["a disclosed note"],
                         mapping_rules=["a mapping rule"])
    text = out.read_text()
    assert "# Test reconciliation" in text
    assert "Office" in text and "Independent" in text
    assert "a disclosed note" in text and "a mapping rule" in text
    assert "GDP (income side)" in text
    assert "2018" not in text          # see docstring: `year` is unused
    assert "zzz" in text                      # unclassified accounts surfaced
